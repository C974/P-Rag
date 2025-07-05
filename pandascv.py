#!/usr/bin/env python3
"""
Palestine Benchmark Analysis Script

This script combines benchmark results from:
1. detailed_benchmark_palestine_opensource - Open source models benchmark data
2. detailed_benchmark_palestine_rag - RAG system with different language models and embeddings

Generates exactly two comprehensive output files:
- Combined CSV with all detailed metrics
- Combined HTML with all detailed metrics and styling
"""

import pandas as pd
import numpy as np
import json
import os
import glob
import seaborn as sns
import matplotlib.pyplot as plt
from datetime import datetime

class PalestineBenchmarkAnalyzer:
    def __init__(self):
        """Initialize the benchmark analyzer"""
        self.rag_folder = "detailed_benchmark_palestine_rag"
        self.opensource_folder = "detailed_benchmark_palestine_opensource"
        
        # Data storage
        self.all_results = []
        
    def load_json_files(self, folder_path: str) -> list:
        """Load all JSON files from a folder"""
        json_files = []
        
        if not os.path.exists(folder_path):
            print(f"Warning: Folder {folder_path} not found!")
            return json_files
        
        # Find all JSON files
        json_pattern = os.path.join(folder_path, "**", "*.json")
        file_paths = glob.glob(json_pattern, recursive=True)
        
        for file_path in file_paths:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Add file info to the data
                    data['source_file'] = os.path.basename(file_path)
                    data['source_folder'] = os.path.basename(folder_path)
                    json_files.append(data)
                    print(f"Loaded: {file_path}")
            except Exception as e:
                print(f"Error loading {file_path}: {e}")
                continue
        
        return json_files
    
    def extract_rag_metrics(self, rag_data: dict) -> dict:
        """Extract detailed metrics from RAG benchmark data"""
        # Basic model info
        metadata = rag_data.get('metadata', {})
        analysis = rag_data.get('analysis', {})
        
        # Extract model name from metadata or derive from folder name
        model_name = metadata.get('model', 'unknown')
        embedding_model = metadata.get('embedding_model', 'unknown')
        
        total_questions = metadata.get('total_questions', 0)
        correct_answers = analysis.get('correct_answers', 0)
        
        metrics = {
            'File_Name': rag_data.get('source_file', 'unknown'),
            'Model_Name': model_name,
            'Embedding_Model': embedding_model,
            'Total_Questions': total_questions,
            'Correct_Answers': correct_answers,
            'Wrong_Answers': total_questions - correct_answers,
            'Overall_Accuracy': analysis.get('overall_accuracy', 0),
        }
        
        # Extract Bloom taxonomy detailed performance
        bloom_stats = analysis.get('bloom_level_stats', {})
        bloom_levels = ['Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create']
        
        for level in bloom_levels:
            if level in bloom_stats:
                stats = bloom_stats[level]
                total = stats.get('total', 0)
                correct = stats.get('correct', 0)
                accuracy = stats.get('accuracy', 0)
                
                metrics[f'{level}_Total_Questions'] = total
                metrics[f'{level}_Correct_Answers'] = correct
                metrics[f'{level}_Wrong_Answers'] = total - correct
                metrics[f'{level}_Accuracy'] = accuracy
            else:
                # If level not found, set to 0
                metrics[f'{level}_Total_Questions'] = 0
                metrics[f'{level}_Correct_Answers'] = 0
                metrics[f'{level}_Wrong_Answers'] = 0
                metrics[f'{level}_Accuracy'] = 0
        
        return metrics
    
    def extract_opensource_metrics(self, opensource_data: dict) -> dict:
        """Extract detailed metrics from open source model benchmark data"""
        # Basic model info  
        metadata = opensource_data.get('metadata', {})
        
        # Extract model name from filename if not in metadata
        source_file = opensource_data.get('source_file', 'unknown')
        model_name = source_file.replace('detailed_benchmark_results_', '').replace('detailed_benchmark_', '').replace('.json', '')
        
        total_questions = metadata.get('total_questions', 0)
        correct_answers = metadata.get('correct_answers', 0)
        wrong_answers = metadata.get('wrong_answers', 0)
        
        metrics = {
            'File_Name': source_file,
            'Model_Name': model_name,
            'Embedding_Model': 'N/A',
            'Total_Questions': total_questions,
            'Correct_Answers': correct_answers,
            'Wrong_Answers': wrong_answers,
            'Overall_Accuracy': metadata.get('overall_accuracy', 0),
        }
        
        # Extract Bloom taxonomy detailed performance
        bloom_analysis = opensource_data.get('bloom_taxonomy_analysis', {})
        bloom_levels = ['Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create']
        
        for level in bloom_levels:
            if level in bloom_analysis:
                stats = bloom_analysis[level]
                total = stats.get('total', 0)
                correct = stats.get('correct', 0)
                accuracy = stats.get('accuracy', 0)
                
                metrics[f'{level}_Total_Questions'] = total
                metrics[f'{level}_Correct_Answers'] = correct
                metrics[f'{level}_Wrong_Answers'] = total - correct
                metrics[f'{level}_Accuracy'] = accuracy
            else:
                # If level not found, set to 0
                metrics[f'{level}_Total_Questions'] = 0
                metrics[f'{level}_Correct_Answers'] = 0
                metrics[f'{level}_Wrong_Answers'] = 0
                metrics[f'{level}_Accuracy'] = 0
        
        return metrics
    
    def load_and_process_data(self):
        """Load and process data from both folders"""
        print("Loading RAG benchmark data...")
        rag_data = self.load_json_files(self.rag_folder)
        
        print("\nLoading Open Source benchmark data...")
        opensource_data = self.load_json_files(self.opensource_folder)
        
        # Process RAG data
        for data in rag_data:
            try:
                metrics = self.extract_rag_metrics(data)
                self.all_results.append(metrics)
            except Exception as e:
                print(f"Error processing RAG data: {e}")
                continue
        
        # Process Open Source data
        for data in opensource_data:
            try:
                metrics = self.extract_opensource_metrics(data)
                self.all_results.append(metrics)
            except Exception as e:
                print(f"Error processing Open Source data: {e}")
                continue
        
        print(f"\nTotal benchmark results loaded: {len(self.all_results)}")
        return len(self.all_results) > 0
    
    def create_dataframe(self):
        """Create and organize the comprehensive dataframe"""
        if not self.all_results:
            print("No data available!")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(self.all_results)
        
        # Define the desired column order
        column_order = [
            'File_Name', 'Model_Name', 'Embedding_Model',
            'Total_Questions', 'Correct_Answers', 'Wrong_Answers', 'Overall_Accuracy'
        ]
        
        # Add Bloom taxonomy columns in order
        bloom_levels = ['Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create']
        for level in bloom_levels:
            column_order.extend([
                f'{level}_Total_Questions',
                f'{level}_Correct_Answers', 
                f'{level}_Wrong_Answers',
                f'{level}_Accuracy'
            ])
        
        # Reorder columns (only include columns that exist)
        existing_columns = [col for col in column_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in existing_columns]
        final_columns = existing_columns + remaining_columns
        
        df = df[final_columns]
        
        # Sort by overall accuracy for better presentation
        df = df.sort_values('Overall_Accuracy', ascending=False).reset_index(drop=True)
        
        return df
    
    def create_heatmap_visualizations(self, df):
        """Create heatmap visualizations using seaborn with model type distinction"""
        print("Creating heatmap visualizations...")
        
        try:
            # Set up seaborn style
            sns.set_style("whitegrid")
            cm = sns.light_palette("green", as_cmap=True)
            
            # Prepare data for heatmap - select key metrics
            heatmap_columns = ['Overall_Accuracy', 'Remember_Accuracy', 'Understand_Accuracy', 
                             'Apply_Accuracy', 'Analyze_Accuracy', 'Evaluate_Accuracy', 'Create_Accuracy']
            
            # Filter columns that exist in dataframe
            available_columns = [col for col in heatmap_columns if col in df.columns]
            
            if len(available_columns) > 1:
                # Create heatmap data with model type indicators
                heatmap_data = df[['Model_Name', 'Embedding_Model'] + available_columns].copy()
                
                # Create model labels with type indicators using text symbols
                model_labels = []
                model_types = []
                for _, row in heatmap_data.iterrows():
                    if row['Embedding_Model'] != 'N/A':
                        # RAG model - add RAG prefix
                        label = f"[RAG] {row['Model_Name']}"
                        model_types.append('RAG')
                    else:
                        # Open source model - add OSM prefix  
                        label = f"[OSM] {row['Model_Name']}"
                        model_types.append('OSM')
                    model_labels.append(label)
                
                # Set model labels as index
                heatmap_data['Model_Label'] = model_labels
                heatmap_data['Model_Type'] = model_types
                heatmap_data = heatmap_data.set_index('Model_Label')[available_columns]
                
                # Create the heatmap with larger figure
                plt.figure(figsize=(16, max(12, len(heatmap_data) * 0.8)))
                
                # Create heatmap with custom styling
                ax = sns.heatmap(heatmap_data, 
                                annot=True, 
                                cmap=cm, 
                                fmt='.3f', 
                                cbar_kws={'label': 'Accuracy Score', 'shrink': 0.8},
                                linewidths=2,
                                linecolor='white',
                                square=False,
                                annot_kws={'size': 10, 'weight': 'bold'})
                
                # Customize y-axis labels with colors
                y_labels = ax.get_yticklabels()
                for i, label in enumerate(y_labels):
                    if '[RAG]' in label.get_text():
                        label.set_color('#1976D2')  # Blue for RAG
                        label.set_weight('bold')
                    elif '[OSM]' in label.get_text():
                        label.set_color('#D32F2F')  # Red for Open Source
                        label.set_weight('bold')
                
                # Customize the plot
                plt.title('Palestine Benchmark Results - Model Performance Heatmap\n[RAG] = RAG Models (Blue) | [OSM] = Open Source Models (Red)', 
                         fontsize=18, fontweight='bold', pad=30)
                plt.xlabel('Performance Metrics', fontsize=14, fontweight='bold')
                plt.ylabel('Models (by Type)', fontsize=14, fontweight='bold')
                plt.xticks(rotation=45, ha='right', fontsize=12)
                plt.yticks(rotation=0, ha='right', fontsize=11)
                
                # Add a colored legend box
                from matplotlib.patches import Rectangle
                import matplotlib.patches as mpatches
                
                # Create legend patches
                rag_patch = mpatches.Patch(color='#1976D2', label='[RAG] RAG Models')
                OSM_patch = mpatches.Patch(color='#D32F2F', label='[OSM] Open Source Models')
                
                # Add legend
                plt.legend(handles=[rag_patch, OSM_patch], 
                          loc='upper left', 
                          bbox_to_anchor=(1.02, 1),
                          fontsize=12,
                          frameon=True,
                          fancybox=True,
                          shadow=True)
                
                # Add grid for better readability
                ax.grid(True, alpha=0.3, linestyle='--')
                
                # Adjust layout to prevent cutoff
                plt.tight_layout()
                
                # Save heatmap with high quality
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                heatmap_path = f'combined_benchmark_heatmap_{timestamp}.png'
                plt.savefig(heatmap_path, dpi=300, bbox_inches='tight', facecolor='white', 
                           edgecolor='none', pad_inches=0.3)
                plt.show()
                print(f"Heatmap visualization saved: {heatmap_path}")
                print(f"🔵 RAG Models: {len([x for x in df['Embedding_Model'] if x != 'N/A'])} models")
                print(f"🔴 Open Source Models: {len([x for x in df['Embedding_Model'] if x == 'N/A'])} models")
                
                return heatmap_path
            else:
                print("Not enough data columns for heatmap visualization")
                return None
                
        except Exception as e:
            print(f"Error creating heatmap visualization: {e}")
            return None
    
    def apply_gradient_styling(self, df):
        """Apply seaborn-based gradient styling to DataFrame"""
        try:
            # Create seaborn colormaps
            green_cm = sns.light_palette("green", as_cmap=True)
            blue_cm = sns.light_palette("blue", as_cmap=True)
            red_cm = sns.light_palette("red", as_cmap=True)
            
            # Apply styling with gradients
            styled_df = df.style.set_caption("Palestine Benchmark Results - Comprehensive Analysis")
            
            # Define columns for different color gradients
            accuracy_columns = [col for col in df.columns if 'Accuracy' in col]
            correct_columns = [col for col in df.columns if 'Correct_Answers' in col]
            wrong_columns = [col for col in df.columns if 'Wrong_Answers' in col]
            total_columns = [col for col in df.columns if 'Total_Questions' in col]
            
            # Apply GREEN gradient to accuracy columns
            for col in accuracy_columns:
                if col in df.columns:
                    styled_df = styled_df.background_gradient(subset=[col], cmap=green_cm, vmin=0, vmax=1)
            
            # Apply BLUE gradient to correct answers and total questions
            for col in correct_columns + total_columns:
                if col in df.columns:
                    styled_df = styled_df.background_gradient(subset=[col], cmap=blue_cm)
            
            # Apply RED gradient to wrong answers
            for col in wrong_columns:
                if col in df.columns:
                    styled_df = styled_df.background_gradient(subset=[col], cmap=red_cm)
            
            # Apply text formatting for better readability
            styled_df = styled_df.format({
                col: '{:.3f}' for col in accuracy_columns if col in df.columns
            })
            
            # Set table styles
            styled_df = styled_df.set_table_styles([
                {'selector': 'caption', 
                 'props': [('font-size', '18px'), ('font-weight', 'bold'), ('text-align', 'center'), ('margin-bottom', '20px')]},
                {'selector': 'th', 
                 'props': [('background-color', '#2e7d32'), ('color', 'white'), ('font-weight', 'bold'), 
                          ('padding', '12px'), ('text-align', 'center'), ('border', '1px solid #1b5e20')]},
                {'selector': 'td', 
                 'props': [('padding', '8px'), ('text-align', 'center'), ('border', '1px solid #ddd')]},
                {'selector': 'table', 
                 'props': [('border-collapse', 'collapse'), ('margin', '0 auto'), ('width', '100%'),
                          ('box-shadow', '0 4px 6px rgba(0,0,0,0.1)'), ('border-radius', '10px'),
                          ('overflow', 'hidden')]}
            ])
            
            return styled_df
            
        except Exception as e:
            print(f"Warning: Could not apply gradient styling - {e}")
            return df.style.set_caption("Palestine Benchmark Results - Comprehensive Analysis")

    def generate_summary_statistics(self, df):
        """Generate comprehensive summary statistics"""
        total_models = len(df)
        rag_models = len([x for x in df['Embedding_Model'] if x != 'N/A'])
        opensource_models = len([x for x in df['Embedding_Model'] if x == 'N/A'])
        
        # Overall statistics
        overall_stats = {
            'total_models': total_models,
            'rag_models': rag_models,
            'opensource_models': opensource_models,
            'best_accuracy': df['Overall_Accuracy'].max(),
            'worst_accuracy': df['Overall_Accuracy'].min(),
            'mean_accuracy': df['Overall_Accuracy'].mean(),
            'median_accuracy': df['Overall_Accuracy'].median()
        }
        
        # Model type comparison
        if rag_models > 0 and opensource_models > 0:
            rag_avg = df[df['Embedding_Model'] != 'N/A']['Overall_Accuracy'].mean()
            opensource_avg = df[df['Embedding_Model'] == 'N/A']['Overall_Accuracy'].mean()
            overall_stats['rag_avg_accuracy'] = rag_avg
            overall_stats['opensource_avg_accuracy'] = opensource_avg
            overall_stats['accuracy_difference'] = abs(rag_avg - opensource_avg)
        
        # Top performers
        top_3 = df.head(3)
        overall_stats['top_performers'] = []
        for _, row in top_3.iterrows():
            model_type = 'RAG' if row['Embedding_Model'] != 'N/A' else 'Open Source'
            overall_stats['top_performers'].append({
                'name': row['Model_Name'],
                'type': model_type,
                'accuracy': row['Overall_Accuracy'],
                'embedding': row['Embedding_Model'] if row['Embedding_Model'] != 'N/A' else 'N/A'
            })
        
        # Bloom taxonomy performance
        bloom_levels = ['Remember', 'Understand', 'Apply', 'Analyze', 'Evaluate', 'Create']
        bloom_stats = {}
        for level in bloom_levels:
            accuracy_col = f'{level}_Accuracy'
            if accuracy_col in df.columns:
                bloom_stats[level] = {
                    'mean_accuracy': df[accuracy_col].mean(),
                    'best_accuracy': df[accuracy_col].max(),
                    'worst_accuracy': df[accuracy_col].min()
                }
        overall_stats['bloom_performance'] = bloom_stats
        
        return overall_stats

    def apply_html_styling(self, df):
        """Apply professional styling with gradients to the HTML output"""
        # Generate summary statistics
        summary_stats = self.generate_summary_statistics(df)
        
        # Apply gradient styling using seaborn
        styled_df = self.apply_gradient_styling(df)
        
        # Create enhanced HTML with summary section
        html_string = '''
        <html>
        <head>
            <title>Palestine Benchmark Results - Comprehensive Analysis with Heatmaps</title>
            <style>
                body {
                    font-family: 'Segoe UI', Arial, sans-serif;
                    margin: 20px;
                    background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
                    min-height: 100vh;
                }
                .header {
                    text-align: center;
                    background: linear-gradient(135deg, #2e7d32 0%, #4caf50 100%);
                    color: white;
                    padding: 30px;
                    border-radius: 15px;
                    margin-bottom: 30px;
                    box-shadow: 0 8px 16px rgba(0,0,0,0.1);
                }
                .summary-section {
                    background: linear-gradient(135deg, #fff3e0 0%, #ffe0b2 100%);
                    padding: 25px;
                    border-radius: 15px;
                    margin-bottom: 30px;
                    border-left: 6px solid #ff9800;
                    box-shadow: 0 6px 12px rgba(0,0,0,0.1);
                }
                .summary-grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                    gap: 20px;
                    margin-top: 20px;
                }
                .summary-card {
                    background: white;
                    padding: 20px;
                    border-radius: 10px;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                    border-top: 4px solid #ff9800;
                }
                .summary-title {
                    font-size: 1.2em;
                    font-weight: bold;
                    color: #e65100;
                    margin-bottom: 15px;
                }
                .summary-item {
                    margin: 8px 0;
                    display: flex;
                    justify-content: space-between;
                }
                .summary-label {
                    font-weight: 500;
                    color: #424242;
                }
                .summary-value {
                    font-weight: bold;
                    color: #1976d2;
                }
                .info {
                    background: linear-gradient(135deg, #e8f5e9 0%, #f1f8e9 100%);
                    padding: 20px;
                    border-radius: 10px;
                    margin-bottom: 30px;
                    border-left: 5px solid #4caf50;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.05);
                }
                .metrics-grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 15px;
                    margin-bottom: 30px;
                }
                .metric-card {
                    background: white;
                    padding: 20px;
                    border-radius: 10px;
                    text-align: center;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                    border-top: 4px solid #4caf50;
                }
                .metric-value {
                    font-size: 2em;
                    font-weight: bold;
                    color: #2e7d32;
                }
                .metric-label {
                    color: #666;
                    margin-top: 5px;
                }
                .table-container {
                    background: white;
                    border-radius: 15px;
                    padding: 20px;
                    box-shadow: 0 8px 16px rgba(0,0,0,0.1);
                    overflow-x: auto;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 0;
                }
                .gradient-note {
                    background: #fff3e0;
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                    border-left: 4px solid #ff9800;
                    font-style: italic;
                }
                .top-performer {
                    background: #e8f5e9;
                    padding: 10px;
                    margin: 5px 0;
                    border-radius: 5px;
                    border-left: 3px solid #4caf50;
                }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🎯 Palestine Benchmark Results</h1>
                <h2>Comprehensive Analysis with Heat Map Gradients</h2>
                <p>RAG vs Open Source Models Performance Comparison</p>
            </div>
            
            <div class="summary-section">
                <h2>📊 Executive Summary</h2>
                <div class="summary-grid">
                    <div class="summary-card">
                        <div class="summary-title">📈 Overall Performance</div>
                        <div class="summary-item">
                            <span class="summary-label">Total Models:</span>
                            <span class="summary-value">''' + str(summary_stats['total_models']) + '''</span>
                        </div>
                        <div class="summary-item">
                            <span class="summary-label">Best Accuracy:</span>
                            <span class="summary-value">''' + f"{summary_stats['best_accuracy']:.3f}" + '''</span>
                        </div>
                        <div class="summary-item">
                            <span class="summary-label">Mean Accuracy:</span>
                            <span class="summary-value">''' + f"{summary_stats['mean_accuracy']:.3f}" + '''</span>
                        </div>
                        <div class="summary-item">
                            <span class="summary-label">Median Accuracy:</span>
                            <span class="summary-value">''' + f"{summary_stats['median_accuracy']:.3f}" + '''</span>
                        </div>
                    </div>
                    
                    <div class="summary-card">
                        <div class="summary-title">🔄 Model Type Breakdown</div>
                        <div class="summary-item">
                            <span class="summary-label">RAG Models:</span>
                            <span class="summary-value">''' + str(summary_stats['rag_models']) + '''</span>
                        </div>
                        <div class="summary-item">
                            <span class="summary-label">Open Source Models:</span>
                            <span class="summary-value">''' + str(summary_stats['opensource_models']) + '''</span>
                        </div>'''
        
        # Add model comparison if both types exist
        if 'rag_avg_accuracy' in summary_stats:
            html_string += f'''
                        <div class="summary-item">
                            <span class="summary-label">RAG Avg Accuracy:</span>
                            <span class="summary-value">{summary_stats['rag_avg_accuracy']:.3f}</span>
                        </div>
                        <div class="summary-item">
                            <span class="summary-label">OSM Avg Accuracy:</span>
                            <span class="summary-value">{summary_stats['opensource_avg_accuracy']:.3f}</span>
                        </div>'''
        
        html_string += '''
                    </div>
                    
                    <div class="summary-card">
                        <div class="summary-title">🏆 Top Performers</div>'''
        
        # Add top performers
        for i, performer in enumerate(summary_stats['top_performers'], 1):
            html_string += f'''
                        <div class="top-performer">
                            <strong>{i}. {performer['name']}</strong><br>
                            <small>{performer['type']} | Accuracy: {performer['accuracy']:.3f}</small>
                        </div>'''
        
        html_string += '''
                    </div>
                </div>
            </div>
            
            <div class="info">
                <div class="metrics-grid">
                    <div class="metric-card">
                        <div class="metric-value">''' + str(len(df)) + '''</div>
                        <div class="metric-label">Total Models</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">''' + str(len([x for x in df['Embedding_Model'] if x != 'N/A'])) + '''</div>
                        <div class="metric-label">RAG Models</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">''' + str(len([x for x in df['Embedding_Model'] if x == 'N/A'])) + '''</div>
                        <div class="metric-label">Open Source Models</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">''' + f"{df['Overall_Accuracy'].max():.3f}" + '''</div>
                        <div class="metric-label">Best Accuracy</div>
                    </div>
                </div>
                <p><strong>Generated:</strong> ''' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '''</p>
            </div>
            
            <div class="gradient-note">
                <strong>🌈 Heat Map Legend:</strong> 
                <br>🟢 <strong>Green gradients</strong> = Performance levels (accuracy scores) - darker green = better performance
                <br>🔵 <strong>Blue gradients</strong> = Positive metrics (total questions, correct answers)
                <br>🔴 <strong>Red gradients</strong> = Wrong answers - darker red = more wrong answers
                <br>🔵 <strong>[RAG]</strong> = RAG Models | 🔴 <strong>[OSM]</strong> = Open Source Models
            </div>
            
            <div class="table-container">
                <h3>📋 Detailed Results Table</h3>
        '''
        
        # Convert styled DataFrame to HTML
        table_html = styled_df.to_html(escape=False, table_id='benchmark_table')
        
        html_string += table_html + '''
            </div>
        </body>
        </html>
        '''
        
        return html_string
    
    def export_results(self, df):
        """Export results in multiple formats with heatmap gradients"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. Export regular CSV
        csv_filename = f'combined_benchmark_results_all_models_{timestamp}.csv'
        df.to_csv(csv_filename, index=False)
        print(f"✅ Standard CSV saved: {csv_filename}")
        
        # 2. Create styled Excel-compatible file using pandas ExcelWriter
        styled_excel_filename = f'combined_benchmark_results_all_models_styled_{timestamp}.xlsx'
        try:
            with pd.ExcelWriter(styled_excel_filename, engine='openpyxl') as writer:
                # Write the dataframe
                df.to_excel(writer, sheet_name='Benchmark Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Benchmark Results']
                
                # Import necessary openpyxl modules for styling
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.formatting.rule import ColorScaleRule
                
                # Apply header styling
                header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in worksheet[1]:  # First row (headers)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Apply conditional formatting (heatmap effect)
                # Green gradient for accuracy columns
                accuracy_cols = [i for i, col in enumerate(df.columns, 1) if 'Accuracy' in col]
                for col_idx in accuracy_cols:
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    range_str = f'{col_letter}2:{col_letter}{len(df)+1}'
                    rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='C8E6C9',
                                        end_type='max', end_color='2E7D32')
                    worksheet.conditional_formatting.add(range_str, rule)
                
                # Red gradient for wrong answers
                wrong_cols = [i for i, col in enumerate(df.columns, 1) if 'Wrong_Answers' in col]
                for col_idx in wrong_cols:
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    range_str = f'{col_letter}2:{col_letter}{len(df)+1}'
                    rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='FFCDD2',
                                        end_type='max', end_color='D32F2F')
                    worksheet.conditional_formatting.add(range_str, rule)
                
                # Blue gradient for correct answers and totals
                positive_cols = [i for i, col in enumerate(df.columns, 1) 
                               if 'Correct_Answers' in col or 'Total_Questions' in col]
                for col_idx in positive_cols:
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    range_str = f'{col_letter}2:{col_letter}{len(df)+1}'
                    rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='BBDEFB',
                                        end_type='max', end_color='1976D2')
                    worksheet.conditional_formatting.add(range_str, rule)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ Styled Excel file saved: {styled_excel_filename}")
        
        except ImportError:
            print("⚠️  openpyxl not available, creating HTML styled CSV instead")
            # Fallback to HTML styled CSV
            styled_df = self.apply_gradient_styling(df)
            styled_csv_filename = f'combined_benchmark_results_all_models_styled_{timestamp}.html'
            
            styled_html_simple = f'''
            <html>
            <head>
                <meta charset="utf-8">
                <title>Palestine Benchmark Results - Styled</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                    th {{ background-color: #2e7d32; color: white; font-weight: bold; }}
                </style>
            </head>
            <body>
                <h2>Palestine Benchmark Results - Heat Map Styled</h2>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                {styled_df.to_html()}
            </body>
            </html>
            '''
            
            with open(styled_csv_filename, 'w', encoding='utf-8') as f:
                f.write(styled_html_simple)
            styled_excel_filename = styled_csv_filename
            print(f"✅ Styled CSV (HTML format) saved: {styled_csv_filename}")
        
        # 3. Create heatmap visualization
        heatmap_file = self.create_heatmap_visualizations(df)
        
        # 4. Export comprehensive styled HTML
        html_content = self.apply_html_styling(df)
        html_filename = f'combined_benchmark_results_all_models_{timestamp}.html'
        
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"✅ Comprehensive styled HTML saved: {html_filename}")
        
        return csv_filename, html_filename, styled_excel_filename, heatmap_file
    
    def display_summary(self, df):
        """Display a summary of the results"""
        print("\n" + "="*80)
        print("PALESTINE BENCHMARK ANALYSIS - COMPREHENSIVE RESULTS")
        print("="*80)
        
        total_models = len(df)
        rag_models = len([x for x in df['Embedding_Model'] if x != 'N/A'])
        opensource_models = len([x for x in df['Embedding_Model'] if x == 'N/A'])
        
        print(f"\nDataset Summary:")
        print(f"  Total models analyzed: {total_models}")
        print(f"  RAG models: {rag_models}")
        print(f"  Open Source models: {opensource_models}")
        
        print(f"\nTop 5 Models by Overall Accuracy:")
        print("-" * 50)
        top_5 = df.head(5)[['Model_Name', 'Embedding_Model', 'Overall_Accuracy']]
        for i, (_, row) in enumerate(top_5.iterrows(), 1):
            embedding = row['Embedding_Model'] if row['Embedding_Model'] != 'N/A' else 'Open Source'
            print(f"  {i}. {row['Model_Name']} ({embedding}): {row['Overall_Accuracy']:.3f}")
        
        if rag_models > 0 and opensource_models > 0:
            rag_avg = df[df['Embedding_Model'] != 'N/A']['Overall_Accuracy'].mean()
            opensource_avg = df[df['Embedding_Model'] == 'N/A']['Overall_Accuracy'].mean()
            print(f"\nModel Type Comparison:")
            print(f"  RAG models average accuracy: {rag_avg:.3f}")
            print(f"  Open Source models average accuracy: {opensource_avg:.3f}")
        
        print("\n" + "="*80)
    
    def run_analysis(self):
        """Run the complete analysis pipeline"""
        print("🚀 Starting Palestine Benchmark Analysis with Heat Map Gradients...")
        print("This will generate comprehensive files with seaborn-based heat map visualizations.")
        
        # Load and process data
        if not self.load_and_process_data():
            print("Failed to load data. Exiting.")
            return
        
        # Create comprehensive dataframe
        print("\n📊 Creating comprehensive dataset...")
        df = self.create_dataframe()
        if df is None:
            print("Failed to create dataset. Exiting.")
            return
        
        # Export results with heat maps
        print("\n🎨 Exporting comprehensive results with heat map gradients...")
        export_results = self.export_results(df)
        csv_file, html_file, styled_csv_file, heatmap_file = export_results
        
        # Display summary
        self.display_summary(df)
        
        print(f"\n🎉 Analysis complete!")
        print(f"\n📁 Generated files:")
        print(f"   1. {csv_file} - Standard CSV for data analysis")
        print(f"   2. {styled_csv_file} - Heat map styled Excel/CSV file")
        print(f"   3. {html_file} - Comprehensive HTML with summary and gradients")
        if heatmap_file:
            print(f"   4. {heatmap_file} - Heat map visualization (PNG) with [RAG]/[OSM] markers")
        
        print(f"\n💡 Tips:")
        print(f"   • Open the styled Excel file to see heat map gradients in Excel")
        print(f"   • The HTML file includes executive summary and best viewing experience")
        print(f"   • 🟢 Green gradients = performance levels (darker = better)")
        print(f"   • 🔵 Blue gradients = positive metrics (questions, correct answers)")
        print(f"   • 🔴 Red gradients = wrong answers (darker = more errors)")
        print(f"   • PNG heatmap: [RAG] = RAG models (blue text), [OSM] = Open source models (red text)")


def main():
    """Main function"""
    analyzer = PalestineBenchmarkAnalyzer()
    analyzer.run_analysis()


if __name__ == "__main__":
    main()
